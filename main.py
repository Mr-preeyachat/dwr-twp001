import os
import io
import re
import json
from datetime import datetime
from calendar import monthrange
from urllib.parse import quote
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
import pandas as pd
from shapely.geometry import Point
import geopandas as gpd

# =============================================================================
# สร้าง FastAPI Instance เพียงครั้งเดียว
# =============================================================================
app = FastAPI(
    title="ระบบตรวจสอบแบบฟอร์มโครงการ & เช็คพิกัดเชิงพื้นที่ (Python FastAPI)",
    description="Backend สำหรับตรวจสอบไฟล์ Excel ตามหลักเกณฑ์ process_check_C และตรวจสอบพิกัดเชิงพื้นที่ (process_check2)",
    version="2.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# -----------------------------------------------------------------------------
# 1. โหลดข้อมูลอ้างอิง (location_ref_2.xlsx)
# -----------------------------------------------------------------------------
excel_ref_path = os.path.join(BASE_DIR, "assets", "location_ref_2.xlsx")
if not os.path.exists(excel_ref_path):
    excel_ref_path = os.path.join(BASE_DIR, "location_ref_2.xlsx")

Province_dict = set()
amphoe_dict = {}
tumbon_dict = {}
basin_main_dict = set()
basin_sub_dict = {}
plan_dict = set()
activ_dict = {}
characteristics_dict = {}

if os.path.exists(excel_ref_path):
    xls_ref = pd.ExcelFile(excel_ref_path)

    # 1) province_amphoe
    df_pa = pd.read_excel(xls_ref, sheet_name='province_amphoe', header=None, skiprows=1)
    for _, r in df_pa.iterrows():
        items = r.dropna().astype(str).str.strip().tolist()
        if items:
            prov = items[0]
            Province_dict.add(prov)
            amphoe_dict[prov] = items[1:]

    # 2) amphoe_tumbon
    df_at = pd.read_excel(xls_ref, sheet_name='amphoe_tumbon', header=None, skiprows=1)
    for _, r in df_at.iterrows():
        items = r.dropna().astype(str).str.strip().tolist()
        if items:
            amp = items[0]
            tumbon_dict[amp] = items[1:]

    # 3) MainRiver_PrincipalRiver
    df_river = pd.read_excel(xls_ref, sheet_name='MainRiver_PrincipalRiver', header=None, skiprows=1)
    for _, r in df_river.iterrows():
        items = r.dropna().astype(str).str.strip().tolist()
        if items:
            main_r = items[0]
            basin_main_dict.add(main_r)
            basin_sub_dict[main_r] = items[1:]

    # 4) Plan_Activities
    df_plan = pd.read_excel(xls_ref, sheet_name='Plan_Activities', header=None, skiprows=1)
    for _, r in df_plan.iterrows():
        items = r.dropna().astype(str).str.strip().tolist()
        if items:
            main_plan = items[0]
            plan_dict.add(main_plan)
            activ_dict[main_plan] = items[1:]

    # 5) Activities_Characteristics
    df_acti = pd.read_excel(xls_ref, sheet_name='Activities_Characteristics', header=None, skiprows=1)
    for _, r in df_acti.iterrows():
        items = r.dropna().astype(str).str.strip().tolist()
        if items:
            charac = items[0]
            characteristics_dict[charac] = items[1:]

# -----------------------------------------------------------------------------
# 2. โหลดไฟล์แผนที่เชิงพื้นที่ (GeoPandas - GeoPackage)
# -----------------------------------------------------------------------------
shp_tambon = os.path.join(BASE_DIR, "assets", "AreaDWR_5.gpkg")
shp_basin = os.path.join(BASE_DIR, "assets", "SB_ONWR_2m.gpkg")
shp_AB  = os.path.join(BASE_DIR, "assets", "Area Based.gpkg")

gdf_tambon = None
gdf_basin = None
gdf_AB = None
sindex_tambon = None
sindex_basin = None
sindex_AB = None

# --- โหลดไฟล์ขอบเขตตำบล ---
if os.path.exists(shp_tambon):
    try:
        # โหลดไฟล์และแปลงระบบพิกัดเป็น EPSG:4326
        gdf_tambon = gpd.read_file(shp_tambon).to_crs(epsg=4326)
        
        if not gdf_tambon.empty:
            sindex_tambon = gdf_tambon.sindex
            print("✅ โหลดไฟล์ขอบเขตตำบลสำเร็จ!")
        else:
            print("⚠️ ไฟล์ขอบเขตตำบลไม่มีข้อมูล (Empty GeoDataFrame)")
            gdf_tambon = None
    except Exception as e:
        print(f"⚠️ ไม่สามารถโหลดไฟล์ขอบเขตตำบลได้: {e}")
else:
    print(f"⚠️ ไม่พบไฟล์ขอบเขตตำบลที่พาธ: {shp_tambon}")

# --- โหลดไฟล์ขอบเขตลุ่มน้ำ ---
if os.path.exists(shp_basin):
    try:
        # โหลดไฟล์และแปลงระบบพิกัดเป็น EPSG:4326
        gdf_basin = gpd.read_file(shp_basin).to_crs(epsg=4326)
        
        if not gdf_basin.empty:
            sindex_basin = gdf_basin.sindex
            print("✅ โหลดไฟล์ขอบเขตลุ่มน้ำสำเร็จ!")
        else:
            print("⚠️ ไฟล์ขอบเขตลุ่มน้ำไม่มีข้อมูล (Empty GeoDataFrame)")
            gdf_basin = None
    except Exception as e:
        print(f"⚠️ ไม่สามารถโหลดไฟล์ขอบเขตลุ่มน้ำได้: {e}")
else:
    print(f"⚠️ ไม่พบไฟล์ขอบเขตลุ่มน้ำที่พาธ: {shp_basin}")

# --- โหลดไฟล์ขอบเขตลุ่มน้ำ ---
if os.path.exists(shp_AB):
    try:
        # โหลดไฟล์และแปลงระบบพิกัดเป็น EPSG:4326
        gdf_AB = gpd.read_file(shp_AB).to_crs(epsg=4326)
        
        if not gdf_AB.empty:
            sindex_AB = gdf_AB.sindex
            print("✅ โหลดไฟล์ขอบเขตABสำเร็จ!")
        else:
            print("⚠️ ไฟล์ขอบเขตABไม่มีข้อมูล (Empty GeoDataFrame)")
            gdf_AB = None
    except Exception as e:
        print(f"⚠️ ไม่สามารถโหลดไฟล์ขอบเขตABได้: {e}")
else:
    print(f"⚠️ ไม่พบไฟล์ขอบเขตลุ่มน้ำที่พาธ: {shp_AB}")

def lookup_spatial_point(lat: float, lon: float):
    """ค้นหาข้อมูลตำบล อำเภอ จังหวัด ลุ่มน้ำหลัก และลุ่มน้ำสาขาจากพิกัด Lat/Long"""
    P_NAME = A_NAME = T_NAME = "ไม่พบ"
    BM_NAME = BS_NAME = "ไม่พบ"
    AB_CODE = AB_NAME = AB_TYPE = "ไม่พบ"

    try:
        pt = Point(float(lon), float(lat))

        if gdf_tambon is not None and sindex_tambon is not None:
            idx_possible = list(sindex_tambon.intersection(pt.bounds))
            for idx in idx_possible:
                if gdf_tambon.geometry.iloc[idx].contains(pt):
                    P_NAME = str(gdf_tambon.iloc[idx, 4]).strip()
                    A_NAME = str(gdf_tambon.iloc[idx, 3]).strip()
                    T_NAME = str(gdf_tambon.iloc[idx, 2]).strip()
                    break

        if gdf_basin is not None and sindex_basin is not None:
            idx_basin = list(sindex_basin.intersection(pt.bounds))
            for idx in idx_basin:
                if gdf_basin.geometry.iloc[idx].contains(pt):
                    BM_NAME = str(gdf_basin.iloc[idx, 2]).strip()
                    BS_NAME = str(gdf_basin.iloc[idx, 1]).strip()
                    break

        if gdf_AB is not None and sindex_AB is not None:
            idx_AB = list(sindex_AB.intersection(pt.bounds))
            for idx in idx_AB:
                if gdf_AB.geometry.iloc[idx].contains(pt):
                    AB_CODE = str(gdf_AB.iloc[idx, 1]).strip()
                    AB_NAME = str(gdf_AB.iloc[idx, 0]).strip()
                    AB_TYPE = str(gdf_AB.iloc[idx, 2]).strip()
                    break

    except Exception as e:
        print("Spatial lookup error:", e)

    return {
        "province": P_NAME,
        "amphoe": A_NAME,
        "tambon": T_NAME,
        "main_basin": BM_NAME,
        "sub_basin": BS_NAME,
        "code_AB": AB_CODE,
        "name_AB": AB_NAME,
        "type_AB": AB_TYPE
    }

# -----------------------------------------------------------------------------
# 3. ฟังก์ชันคำนวณและตรวจสอบข้อความ (Helper Functions)
# -----------------------------------------------------------------------------
def levenshtein_distance(s1, s2):
    if len(s1) < len(s2):
        return levenshtein_distance(s2, s1)
    if len(s2) == 0:
        return len(s1)
    previous_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        current_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = previous_row[j + 1] + 1
            deletions = current_row[j] + 1
            substitutions = previous_row[j] + (c1 != c2)
            current_row.append(min(insertions, deletions, substitutions))
        previous_row = current_row
    return previous_row[-1]

def check_project_name_error(text):
    if not text or not str(text).strip():
        return "ผ่าน"
        
    txt = str(text).strip()
    ans = ""
    ans4 = ""
    ans2 = ""
    ans3 = ""
    
    count = 0
    if "ส่งน้ำ" in txt: count += 1
    if "สูบน้ำ" in txt: count += 1
    if "กระจายน้ำ" in txt: count += 1
    if count > 1:
        ans = "ระบบกระจายน้ำ ซ้ำ"

    count2 = 0
    if "ค่าใช้จ่ายในการ" not in txt and "ค่าจ้าง" not in txt:
        if "ก่อสร้าง" in txt: count2 += 1
        if "ปรับปรุง" in txt: count2 += 1
        if "ซ่อมแซม" in txt: count2 += 1
        if "อนุรักษ์ฟื้นฟู" in txt: count2 += 1
        if count2 > 1:
            ans4 = "กิจกรรม ซ้ำ"
        
    if "ระบบกระจายน้ำเพื่อผลิตน้ำสะอาด" in txt or "ระบบกระจายน้ำสะอาด" in txt:
        ans3 = "ใช้คำว่า ระบบกระจายน้ำสะอาด"
        
    keywords = ["ก่อสร้าง", "ปรับปรุงเพิ่มประสิทธิภาพ", "อนุรักษ์ฟื้นฟู", "ซ่อมแซม", 
                "ระบบกระจายน้ำ", "พลังงานแสงอาทิตย์", "สนับสนุน", "อุปโภค", "บริโภค"]
    typo_list = []

    for kw in keywords:
        if kw in txt:
            continue  # ถ้ามีคำถูกเป๊ะๆ ในข้อความแล้ว ให้ข้าม
        
        len_kw = len(kw)
        max_allowed = 1 if len_kw <= 5 else (2 if len_kw <= 10 else 3)
        min_dist = 9999
        
        # วนลูปตัด chunk ที่มีความยาวต่างกัน (เผื่อพิมพ์ขาดหรือพิมพ์เกิน)
        for window_size in range(max(1, len_kw - max_allowed), len_kw + max_allowed + 1):
            for i in range(0, len(txt) - window_size + 1):
                chunk = txt[i : i + window_size]
                dist = levenshtein_distance(chunk, kw)
                if dist < min_dist:
                    min_dist = dist

        if 0 < min_dist <= max_allowed:
            typo_list.append(kw)
            
    if typo_list:
        ans2 = f"คำผิด: {', '.join(typo_list)}"
        
    errors = [a for a in [ans, ans2, ans3, ans4] if a] 
    return " ".join(errors) if errors else "ผ่าน"

def parse_to_mmyyyy(cell_value):
    if cell_value is None:
        return None

    if isinstance(cell_value, datetime):
        month = f"{cell_value.month:02d}"
        year = cell_value.year
        if 1900 <= year <= 1999:
            year += 600
        elif year < 2200:
            year += 543
        return f"{month}-{year}"

    val_str = str(cell_value).strip()
    val_str = re.sub(r"[/.\s]+", "-", val_str)
    parts = val_str.split("-")

    try:
        if len(parts) == 2:
            m, y = int(parts[0]), int(parts[1])
            if m > 12: m, y = y, m
            if y < 100: y += 2500
            elif 1900 <= y <= 1999: y += 600
            if 1 <= m <= 12 and y > 1000:
                return f"{m:02d}-{y}"
        elif len(parts) == 3:
            if len(parts[0]) >= 4 or (len(parts[0]) == 2 and int(parts[0]) > 31):
                y, m = int(parts[0]), int(parts[1])
            else:
                m, y = int(parts[1]), int(parts[2])
            if y < 100: y += 2500
            elif 1900 <= y <= 1999: y += 600
            if 1 <= m <= 12 and y > 1000:
                return f"{m:02d}-{y}"
    except ValueError:
        return None
    return None

def calculate_rounded_days(start_mmyyyy, end_mmyyyy):
    try:
        m_start, y_start = map(int, start_mmyyyy.split("-"))
        m_end, y_end = map(int, end_mmyyyy.split("-"))

        y_start_ad = y_start - 543 if y_start > 2200 else y_start
        y_end_ad = y_end - 543 if y_end > 2200 else y_end

        dt_start = datetime(y_start_ad, m_start, 1)
        _, last_day = monthrange(y_end_ad, m_end)
        dt_end = datetime(y_end_ad, m_end, last_day)

        if dt_start > dt_end:
            return None

        total_days = (dt_end - dt_start).days + 1
        rounded_days = (total_days // 10) * 10
        return rounded_days
    except Exception:
        return None

def col_letter_to_index(col_letter):
    num = 0
    for char in col_letter.upper():
        num = num * 26 + (ord(char) - ord('A') + 1)
    return num - 1

# -----------------------------------------------------------------------------
# 4. กระบวนการตรวจสอบหลัก (process_check_C)
# -----------------------------------------------------------------------------
def process_check_C_logic(wb):
    ws = wb.active

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    green_font = Font(color="006600")
    red_font = Font(color="CC0000", bold=True)

    name_read_idx = col_letter_to_index("E")   # ชื่อโครงการ
    name_write_idx = col_letter_to_index("C")  # เขียนคำตอบรายงานผล
    check_idx = col_letter_to_index("B")
    check2_idx = col_letter_to_index("D")

    t_idx = col_letter_to_index("H")   # ตำบล
    a_idx = col_letter_to_index("I")   # อำเภอ
    p_idx = col_letter_to_index("J")   # จังหวัด
    bm_idx = col_letter_to_index("K")  # ลุ่มน้ำหลัก
    bs_idx = col_letter_to_index("L")  # ลุ่มน้ำสาขา

    lat_idx = col_letter_to_index("M") # Latitude
    lon_idx = col_letter_to_index("N") # Longitude

    Plan_idx = col_letter_to_index("V")   # แผนงาน
    Activ_idx = col_letter_to_index("T")  # กิจกรรม
    Charac_idx = col_letter_to_index("U") # ลักษณะเฉพาะ

    # KPI ONWR
    KPI_Y_idx = col_letter_to_index("Y"); KPI_Z_idx = col_letter_to_index("Z")
    KPI_AB_idx = col_letter_to_index("AB")
    KPI_AR_idx = col_letter_to_index("AR"); KPI_AS_idx = col_letter_to_index("AS")
    KPI_AU_idx = col_letter_to_index("AU"); KPI_AV_idx = col_letter_to_index("AV"); KPI_AW_idx = col_letter_to_index("AW"); KPI_AX_idx = col_letter_to_index("AX")
    KPI_BG_idx = col_letter_to_index("BG"); KPI_BI_idx = col_letter_to_index("BI")
    KPI_BT_idx = col_letter_to_index("BT"); KPI_BU_idx = col_letter_to_index("BU")
    KPI_BZ_idx = col_letter_to_index("BZ"); KPI_CA_idx = col_letter_to_index("CA"); KPI_CB_idx = col_letter_to_index("CB")
    KPI_CM_idx = col_letter_to_index("CM"); KPI_CN_idx = col_letter_to_index("CN")
    KPI_CO_idx = col_letter_to_index("CO"); KPI_CP_idx = col_letter_to_index("CP")
    KPI_CQ_idx = col_letter_to_index("CQ"); KPI_CR_idx = col_letter_to_index("CR")
    KPI_CS_idx = col_letter_to_index("CS"); KPI_CT_idx = col_letter_to_index("CT")
    KPI_CU_idx = col_letter_to_index("CU"); KPI_CV_idx = col_letter_to_index("CV")
    KPI_CW_idx = col_letter_to_index("CW")
    KPI_CY_idx = col_letter_to_index("CY"); KPI_CZ_idx = col_letter_to_index("CZ")
    KPI_DA_idx = col_letter_to_index("DA"); KPI_DB_idx = col_letter_to_index("DB")
    KPI_DC_idx = col_letter_to_index("DC"); KPI_DD_idx = col_letter_to_index("DD")
    KPI_DE_idx = col_letter_to_index("DE")
    KPI_DH_idx = col_letter_to_index("DH"); KPI_DI_idx = col_letter_to_index("DI"); KPI_DJ_idx = col_letter_to_index("DJ"); KPI_DK_idx = col_letter_to_index("DK"); KPI_DL_idx = col_letter_to_index("DL")

    # KPI DWR
    KPI_DWR_EO_idx = col_letter_to_index("EO"); KPI_DWR_EP_idx = col_letter_to_index("EP"); KPI_DWR_EQ_idx = col_letter_to_index("EQ"); KPI_DWR_ER_idx = col_letter_to_index("ER"); KPI_DWR_ES_idx = col_letter_to_index("ES")

    # Attachments
    Attach_DM_idx = col_letter_to_index("DM"); Attach_DN_idx = col_letter_to_index("DN"); Attach_DO_idx = col_letter_to_index("DO"); Attach_DP_idx = col_letter_to_index("DP")

    # Dates
    day_st_idx = col_letter_to_index("DQ")
    day_en_idx = col_letter_to_index("DR")
    day_n_idx = col_letter_to_index("DS")

    total_rows = 0
    passed_rows = 0
    error_rows = 0

    if ws.max_row < 9:
        return total_rows, passed_rows, error_rows

    for row in ws.iter_rows(min_row=9):
        cell_NR = row[name_read_idx]
        cell_NW = row[name_write_idx]
        cell_check = row[check_idx]
        cell_check2 = row[check2_idx]

        cell_LAT = row[lat_idx]; cell_LONG = row[lon_idx]
        cell_P = row[p_idx]; cell_A = row[a_idx]; cell_T = row[t_idx]
        cell_BM = row[bm_idx]; cell_BS = row[bs_idx]

        cell_Plan = row[Plan_idx]; cell_Activ = row[Activ_idx]; cell_Charac = row[Charac_idx]

        cell_KPI_Y = row[KPI_Y_idx]; cell_KPI_Z = row[KPI_Z_idx]
        cell_KPI_AB = row[KPI_AB_idx]
        cell_KPI_AR = row[KPI_AR_idx]; cell_KPI_AS = row[KPI_AS_idx]
        cell_KPI_AU = row[KPI_AU_idx]; cell_KPI_AV = row[KPI_AV_idx]; cell_KPI_AW = row[KPI_AW_idx]; cell_KPI_AX = row[KPI_AX_idx]
        cell_KPI_BG = row[KPI_BG_idx]; cell_KPI_BI = row[KPI_BI_idx]
        cell_KPI_BT = row[KPI_BT_idx]; cell_KPI_BU = row[KPI_BU_idx]
        cell_KPI_BZ = row[KPI_BZ_idx]; cell_KPI_CA = row[KPI_CA_idx]; cell_KPI_CB = row[KPI_CB_idx]
        cell_KPI_CM = row[KPI_CM_idx]; cell_KPI_CN = row[KPI_CN_idx]
        cell_KPI_CO = row[KPI_CO_idx]; cell_KPI_CP = row[KPI_CP_idx]
        cell_KPI_CQ = row[KPI_CQ_idx]; cell_KPI_CR = row[KPI_CR_idx]
        cell_KPI_CS = row[KPI_CS_idx]; cell_KPI_CT = row[KPI_CT_idx]
        cell_KPI_CU = row[KPI_CU_idx]; cell_KPI_CV = row[KPI_CV_idx]
        cell_KPI_CW = row[KPI_CW_idx]
        cell_KPI_CY = row[KPI_CY_idx]; cell_KPI_CZ = row[KPI_CZ_idx]
        cell_KPI_DA = row[KPI_DA_idx]; cell_KPI_DB = row[KPI_DB_idx]
        cell_KPI_DC = row[KPI_DC_idx]; cell_KPI_DD = row[KPI_DD_idx]
        cell_KPI_DE = row[KPI_DE_idx]
        cell_KPI_DH = row[KPI_DH_idx]; cell_KPI_DI = row[KPI_DI_idx]; cell_KPI_DJ = row[KPI_DJ_idx]; cell_KPI_DK = row[KPI_DK_idx]; cell_KPI_DL = row[KPI_DL_idx]

        cell_KPI_DWR_EO = row[KPI_DWR_EO_idx]; cell_KPI_DWR_EP = row[KPI_DWR_EP_idx]; cell_KPI_DWR_EQ = row[KPI_DWR_EQ_idx]; cell_KPI_DWR_ER = row[KPI_DWR_ER_idx]; cell_KPI_DWR_ES = row[KPI_DWR_ES_idx]
        cell_Attach_DM = row[Attach_DM_idx]; cell_Attach_DN = row[Attach_DN_idx]; cell_Attach_DO = row[Attach_DO_idx]; cell_Attach_DP = row[Attach_DP_idx]

        cell_day_st = row[day_st_idx]; cell_day_en = row[day_en_idx]; cell_day_n = row[day_n_idx]

        val_NR = str(cell_NR.value).strip() if cell_NR.value is not None else ""
        if not val_NR and cell_P.value is None and cell_LAT.value is None:
            continue

        total_rows += 1

        value_P = str(cell_P.value).strip() if cell_P.value else None
        value_A = str(cell_A.value).strip() if cell_A.value else None
        value_T = str(cell_T.value).strip() if cell_T.value else None
        value_BM = str(cell_BM.value).strip() if cell_BM.value else None
        value_BS = str(cell_BS.value).strip() if cell_BS.value else None

        value_Plan = str(cell_Plan.value).strip() if cell_Plan.value else None
        value_Activ = str(cell_Activ.value).strip() if cell_Activ.value else None
        value_Charac = str(cell_Charac.value).strip() if cell_Charac.value else None

        # 1. เชคคำถูก ผิด ชื่อโครงการ
        error_report = check_project_name_error(val_NR) if val_NR else ""
        check_err_word = str(error_report).strip() if error_report and error_report != "ผ่าน" else "ผ่าน"

        # 2. เชคพิกัด (รวม Spatial Intersection)
        check_err_latlong = "ผ่าน"
        check_err_latlong_A = ""
        check_err_latlong_B = ""
        if str(cell_P.value).strip() == "ส่วนกลาง":
            if str(cell_LAT.value).strip() != "13.78553" or str(cell_LONG.value).strip() != "100.53915":
                cell_LAT.fill = red_fill; cell_LONG.fill = red_fill
                check_err_latlong_A = "พิกัดส่วนกลาง"
            if str(cell_A.value).strip() != "ส่วนกลาง":
                cell_A.fill = red_fill
                check_err_latlong_A = "ที่ตั้งต้องเป็นส่วนกลาง"
            if str(cell_T.value).strip() != "ส่วนกลาง":
                cell_T.fill = red_fill
                check_err_latlong_A = "ที่ตั้งต้องเป็นส่วนกลาง"
        else:
            lat_val = float(cell_LAT.value) if cell_LAT.value is not None and str(cell_LAT.value).strip() != "" else None
            lon_val = float(cell_LONG.value) if cell_LONG.value is not None and str(cell_LONG.value).strip() != "" else None

            if lat_val is None or lon_val is None:
                cell_LAT.fill = red_fill; cell_LONG.fill = red_fill
                check_err_latlong_B = "ไม่มีพิกัด"
            else:
                # 2.1 Spatial Intersection Check
                spatial_info = lookup_spatial_point(lat_val, lon_val)
                P_NAME = spatial_info["province"]
                A_NAME = spatial_info["amphoe"]
                T_NAME = spatial_info["tambon"]
                BM_NAME = spatial_info["main_basin"]
                BS_NAME = spatial_info["sub_basin"]

                latlong_error = 0
                if str(cell_P.value).strip() == P_NAME: cell_P.font = green_font
                else: cell_P.font = red_font; latlong_error += 1

                if str(cell_A.value).strip() == A_NAME: cell_A.font = green_font
                else: cell_A.font = red_font; latlong_error += 1

                if str(cell_T.value).strip() == T_NAME: cell_T.font = green_font
                else: cell_T.font = red_font; latlong_error += 1

                if str(cell_BM.value).strip() == BM_NAME: cell_BM.font = green_font
                else: cell_BM.font = red_font; latlong_error += 10

                if str(cell_BS.value).strip() == BS_NAME: cell_BS.font = green_font
                else: cell_BS.font = red_font; latlong_error += 10

                if latlong_error >= 1:
                    cell_LAT.fill = yellow_fill; cell_LONG.fill = yellow_fill
                    check_err_latlong_B = "พิกัดอาจไม่ตรงกับพื้นที่"
                else:
                    cell_LAT.fill = green_fill; cell_LONG.fill = green_fill

        if check_err_latlong_A or check_err_latlong_B:
            check_err_latlong = f"พิกัดผิดพลาด: {check_err_latlong_A} {check_err_latlong_B}".strip()

        # 3. ตรวจสอบลำดับชั้นที่ตั้ง (location_ref)
        check_err_loref_num1 = 0
        check_err_loref_num2 = 0

        if not value_P: cell_P.fill = red_fill; check_err_loref_num1 += 1
        elif value_P in Province_dict: cell_P.fill = green_fill
        else: cell_P.fill = yellow_fill; check_err_loref_num2 += 1

        if not value_A: cell_A.fill = red_fill; check_err_loref_num1 += 1
        elif (cell_P.fill == green_fill) and (value_P in amphoe_dict) and any(kw == value_A for kw in amphoe_dict[value_P]):
            cell_A.fill = green_fill
        else: cell_A.fill = yellow_fill; check_err_loref_num2 += 1

        if not value_T: cell_T.fill = red_fill; check_err_loref_num1 += 1
        elif (cell_A.fill == green_fill) and (value_A in tumbon_dict) and any(kw == value_T for kw in tumbon_dict[value_A]):
            cell_T.fill = green_fill
        else: cell_T.fill = yellow_fill; check_err_loref_num2 += 1

        if not value_BM: cell_BM.fill = red_fill; check_err_loref_num1 += 10
        elif value_BM in basin_main_dict: cell_BM.fill = green_fill
        else: cell_BM.fill = yellow_fill; check_err_loref_num2 += 10

        if not value_BS: cell_BS.fill = red_fill; check_err_loref_num1 += 10
        elif (cell_BM.fill == green_fill) and (value_BM in basin_sub_dict) and any(kw == value_BS for kw in basin_sub_dict[value_BM]):
            cell_BS.fill = green_fill
        else: cell_BS.fill = yellow_fill; check_err_loref_num2 += 10

        check_err_loref = "ที่ตั้งพิมพ์ไม่ถูกต้อง" if (check_err_loref_num1 + check_err_loref_num2) > 0 else "ผ่าน"

        # 4. ตรวจสอบแผนงานและกิจกรรม
        check_err_plan_num1 = 0
        check_err_plan_num2 = 0

        if not value_Plan: cell_Plan.fill = red_fill; check_err_plan_num1 += 1
        elif value_Plan in plan_dict: cell_Plan.fill = green_fill
        else: cell_Plan.fill = yellow_fill; check_err_plan_num2 += 1

        if not value_Activ: cell_Activ.fill = red_fill; check_err_plan_num1 += 1
        elif (cell_Plan.fill == green_fill) and (value_Plan in activ_dict) and any(kw == value_Activ for kw in activ_dict[value_Plan]):
            cell_Activ.fill = green_fill
        else: cell_Activ.fill = yellow_fill; check_err_plan_num2 += 1

        if not value_Charac: cell_Charac.fill = red_fill; check_err_plan_num1 += 1
        else:
            lookup_key = f"{value_Plan}_{value_Activ}"
            if (cell_Plan.fill == green_fill) and (cell_Activ.fill == green_fill) and (lookup_key in characteristics_dict):
                if any(kw == value_Charac for kw in characteristics_dict[lookup_key]):
                    cell_Charac.fill = green_fill
                else: cell_Charac.fill = yellow_fill; check_err_plan_num2 += 1
            else: cell_Charac.fill = yellow_fill; check_err_plan_num2 += 1

        check_err_plan = "แผนงาน ประเภท ไม่ถูกต้อง" if (check_err_plan_num1 + check_err_plan_num2) > 0 else "ผ่าน"

        # 5. ตัวชี้วัด สทนช.
        check_err_onwr_num = 0
        p_val = str(cell_Plan.value).strip() if cell_Plan.value else ""
        if p_val == "1.1.1" and (cell_KPI_Y.value is None or cell_KPI_Z.value is None): cell_KPI_Y.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "1.1.2" and cell_KPI_AB.value is None: cell_KPI_AB.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "2.2.2" and (cell_KPI_AR.value is None or cell_KPI_AS.value is None): cell_KPI_AR.fill = yellow_fill; cell_KPI_AS.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val in ["2.3.1", "2.3.3"] and (cell_KPI_AU.value is None or cell_KPI_AV.value is None or cell_KPI_AW.value is None or cell_KPI_AX.value is None):
            cell_KPI_AU.fill = yellow_fill; cell_KPI_AV.fill = yellow_fill; cell_KPI_AW.fill = yellow_fill; cell_KPI_AX.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "3.1.1" and cell_KPI_BG.value is None: cell_KPI_BG.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "3.1.3" and cell_KPI_BI.value is None: cell_KPI_BI.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "3.4.1" and cell_KPI_BT.value is None and cell_KPI_BU.value is None: cell_KPI_BT.fill = yellow_fill; cell_KPI_BU.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "4.2.1" and cell_KPI_BZ.value is None and cell_KPI_CA.value is None: cell_KPI_BZ.fill = yellow_fill; cell_KPI_CA.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "4.2.2" and (cell_KPI_BZ.value is None or cell_KPI_CB.value is None): cell_KPI_BZ.fill = yellow_fill; cell_KPI_CB.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "4.5.1" and cell_KPI_CM.value is None and cell_KPI_CN.value is None: cell_KPI_CM.fill = yellow_fill; cell_KPI_CN.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "4.5.2" and cell_KPI_CO.value is None: cell_KPI_CO.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "4.5.3" and cell_KPI_CP.value is None: cell_KPI_CP.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "5.1.1" and cell_KPI_CQ.value is None and cell_KPI_CR.value is None: cell_KPI_CQ.fill = yellow_fill; cell_KPI_CR.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "5.2.1" and cell_KPI_CS.value is None: cell_KPI_CS.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "5.2.2" and cell_KPI_CT.value is None: cell_KPI_CT.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "5.2.3" and cell_KPI_CU.value is None and cell_KPI_CV.value is None: cell_KPI_CU.fill = yellow_fill; cell_KPI_CV.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "5.2.4" and cell_KPI_CW.value is None: cell_KPI_CW.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "5.3.2" and cell_KPI_CY.value is None: cell_KPI_CY.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "5.3.3" and cell_KPI_CZ.value is None: cell_KPI_CZ.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "5.3.4" and cell_KPI_DA.value is None and cell_KPI_DB.value is None: cell_KPI_DA.fill = yellow_fill; cell_KPI_DB.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "5.3.5" and cell_KPI_DC.value is None and cell_KPI_DD.value is None: cell_KPI_DC.fill = yellow_fill; cell_KPI_DD.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "5.3.6" and cell_KPI_DE.value is None: cell_KPI_DE.fill = yellow_fill; check_err_onwr_num += 1
        elif p_val == "5.3.8" and cell_KPI_DH.value is None and cell_KPI_DI.value is None and cell_KPI_DJ.value is None and cell_KPI_DK.value is None and cell_KPI_DL.value is None:
            cell_KPI_DH.fill = yellow_fill; cell_KPI_DI.fill = yellow_fill; cell_KPI_DJ.fill = yellow_fill; cell_KPI_DK.fill = yellow_fill; cell_KPI_DL.fill = yellow_fill; check_err_onwr_num += 1

        if check_err_onwr_num > 0:
            check_err_onwr = "ตัวชี้วัด สทนช. ไม่ครบ"
            cell_check.fill = yellow_fill
        else: check_err_onwr = "ผ่าน"

        # 6. ตัวชี้วัด ทน.
        check_err_dwr_num = 0
        act_val = str(cell_Activ.value).strip() if cell_Activ.value else ""
        charac_val = str(cell_Charac.value).strip() if cell_Charac.value else ""
        nr_val = str(cell_NR.value).strip() if cell_NR.value else ""

        if p_val in ["2.2.2", "2.3.1", "2.3.3", "4.5.1", "4.5.2", "4.5.3"] and act_val != "บริหารจัดการ":
            if charac_val != "ระบบกระจายน้ำ" and (cell_KPI_DWR_EO.value is None or cell_KPI_DWR_EP.value is None or cell_KPI_DWR_EQ.value is None):
                cell_check2.fill = yellow_fill; cell_KPI_DWR_EO.fill = yellow_fill; cell_KPI_DWR_EP.fill = yellow_fill; cell_KPI_DWR_EQ.fill = yellow_fill; check_err_dwr_num += 1
            elif charac_val == "ระบบกระจายน้ำ" and nr_val.startswith("อนุรักษ์ฟื้นฟู") and (cell_KPI_DWR_EO.value is None or cell_KPI_DWR_EP.value is None or cell_KPI_DWR_EQ.value is None or cell_KPI_DWR_ER.value is None):
                cell_check2.fill = yellow_fill; cell_KPI_DWR_EO.fill = yellow_fill; cell_KPI_DWR_EP.fill = yellow_fill; cell_KPI_DWR_EQ.fill = yellow_fill; cell_KPI_DWR_ER.fill = yellow_fill; check_err_dwr_num += 1
            elif charac_val == "ระบบกระจายน้ำ" and (cell_KPI_DWR_ER.value is None or cell_KPI_DWR_EP.value is None or cell_KPI_DWR_EQ.value is None):
                cell_check2.fill = yellow_fill; cell_KPI_DWR_EP.fill = yellow_fill; cell_KPI_DWR_EQ.fill = yellow_fill; cell_KPI_DWR_ER.fill = yellow_fill; check_err_dwr_num += 1
            elif nr_val.startswith("ปรับปรุง") and "พร้อมระบบกระจายน้ำ" in nr_val and cell_KPI_DWR_EO.value is None:
                cell_check2.fill = yellow_fill; cell_KPI_DWR_EO.fill = yellow_fill; check_err_dwr_num += 1

        if charac_val == "ระบบประปา" and cell_KPI_DWR_EQ.value is None:
            cell_check2.fill = yellow_fill; cell_KPI_DWR_EQ.fill = yellow_fill; check_err_dwr_num += 1

        if p_val == "3.1.3" and cell_KPI_DWR_ES.value is None:
            cell_check2.fill = yellow_fill; cell_KPI_DWR_ES.fill = yellow_fill; check_err_dwr_num += 1

        check_err_dwr = "ตัวชี้วัด ทน. ไม่ครบ" if check_err_dwr_num > 0 else "ผ่าน"

        # 7. ชื่อตามกิจกรรม
        check_err_name_num = 0
        forbidden_starts = ["ค่าจ้างควบคุมงาน", "ค่าใช้จ่าย", "ก่อสร้าง", "อนุรักษ์ฟื้นฟู", "ปรับปรุง", "ซ่อมแซม", "บำรุงรักษา"]
        if nr_val:
            if not any(nr_val.startswith(k) for k in forbidden_starts) or nr_val.startswith("โครงการ"):
                cell_NR.fill = yellow_fill; check_err_name_num += 1

        if charac_val == "ระบบกระจายน้ำ":
            keywords = ["กระจายน้ำ", "ส่งน้ำ", "สูบน้ำ", "เครือข่ายน้ำ", "โครงข่าย"]
            if not any(k in nr_val for k in keywords):
                cell_NR.fill = yellow_fill; check_err_name_num += 1

        if p_val == "2.2.2" and "เพื่อการถ่ายโอนให้องค์กรปกครองส่วนท้องถิ่น" not in nr_val:
            cell_NR.fill = yellow_fill; check_err_name_num += 1
        if p_val == "2.3.3" and "สนับสนุนพื้นที่ปฏิรูปที่ดินเพื่อการเกษตร (สปก.)" not in nr_val:
            cell_NR.fill = yellow_fill; check_err_name_num += 1
        if p_val in ["4.5.1", "4.5.2", "4.5.3"] and "อนุรักษ์ฟื้นฟู" not in nr_val:
            cell_NR.fill = yellow_fill; check_err_name_num += 1
        if p_val in ["5.1.1","5.2.1","5.2.2","5.2.3","5.2.4","5.3.1","5.3.2","5.3.3","5.3.4","5.3.5","5.3.6","5.3.7","5.3.8","5.4.1"] and "ค่าใช้จ่าย" not in nr_val:
            cell_NR.fill = yellow_fill; check_err_name_num += 1

        check_err_name = "ชื่อไม่สอดคล้องกับประเภทงาน" if check_err_name_num > 0 else "ผ่าน"

        # 8. เอกสารแนบ
        check_err_Attach_num = 0
        if act_val == "บริหารจัดการ":
            if cell_Attach_DM.value != 4 or cell_Attach_DN.value != 4:
                cell_Attach_DM.fill = yellow_fill; cell_Attach_DN.fill = yellow_fill; check_err_Attach_num += 1
        else:
            if cell_Attach_DM.value != 4 or cell_Attach_DN.value != 4 or cell_Attach_DO.value != 4 or cell_Attach_DP.value != 4:
                cell_Attach_DM.fill = yellow_fill; cell_Attach_DN.fill = yellow_fill; cell_Attach_DO.fill = yellow_fill; cell_Attach_DP.fill = yellow_fill; check_err_Attach_num += 1

        check_err_Attach = "เอกสารแนบ ไม่พร้อม" if check_err_Attach_num > 0 else "ผ่าน"

        # 9. เชควันและคำนวณวัน
        check_err_day_num = 0
        Day_start = 2570
        allowed_years = (str(Day_start), str(Day_start + 1))

        formatted_date_start = parse_to_mmyyyy(cell_day_st.value)
        formatted_date_end = parse_to_mmyyyy(cell_day_en.value)

        if formatted_date_start and str(formatted_date_start).endswith(allowed_years):
            cell_day_st.value = formatted_date_start
        else:
            cell_day_st.fill = red_fill; cell_day_st.value = formatted_date_start; check_err_day_num += 1

        if formatted_date_end and str(formatted_date_end).endswith(allowed_years):
            cell_day_en.value = formatted_date_end
        else:
            cell_day_en.fill = red_fill; cell_day_en.value = formatted_date_end; check_err_day_num += 10

        if formatted_date_start and formatted_date_end:
            days_result = calculate_rounded_days(formatted_date_start, formatted_date_end)
            if days_result is not None:
                cell_day_n.value = days_result
            else:
                cell_day_n.fill = red_fill; check_err_day_num += 1

        check_err_day = "รูปแบบวันผิดพลาด" if check_err_day_num > 0 else "ผ่าน"

        # 10. รวบรวมเฉพาะข้อความ คำตอบ ลงคอลัมน์ C (cell_NW)
        check_list = [
            check_err_word,
            check_err_latlong,
            check_err_loref,
            check_err_plan,
            check_err_onwr,
            check_err_dwr,
            check_err_name,
            check_err_Attach,
            check_err_day,
        ]

        check_err_errors = [err for err in check_list if err and err != "ผ่าน"]

        if check_err_errors:
            error_rows += 1
            err_text = " | ".join(check_err_errors)
            cell_NW.value = err_text
            cell_NW.font = red_font
            cell_NW.fill = yellow_fill
            cell_NW.comment = Comment(f"❌ รายงานข้อผิดพลาด:\n- " + "\n- ".join(check_err_errors), "System Validator")
        else:
            passed_rows += 1
            cell_NW.value = "ผ่าน"
            cell_NW.font = green_font
            cell_NW.fill = PatternFill(fill_type=None)

    return total_rows, passed_rows, error_rows

# -----------------------------------------------------------------------------
# 5. API Endpoints
# -----------------------------------------------------------------------------
class CoordRequest(BaseModel):
    lat: float
    lon: float

@app.get("/", response_class=HTMLResponse)
async def serve_index():
    index_path = os.path.join(BASE_DIR, "Index.html")
    if os.path.exists(index_path):
        with open(index_path, "r", encoding="utf-8") as f:
            return f.read()
    return "<h1>Index.html not found</h1>"
#----
@app.get("/map", response_class=HTMLResponse)
async def serve_map():
    map_path = os.path.join(BASE_DIR, "index_map.html")
    if os.path.exists(map_path):
        with open(map_path, "r", encoding="utf-8") as f:
            return f.read()
    return "<h1>index_map.html not found</h1>"
#----
@app.post("/api/check_coords")
async def check_coordinates_api(req: CoordRequest):
    result = lookup_spatial_point(req.lat, req.lon)
    return {"success": True, "data": result}

@app.post("/api/verify")
async def verify_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="กรุณาอัปโหลดไฟล์ Excel (.xlsx หรือ .xls) เท่านั้น")

    try:
        contents = await file.read()
        excel_stream = io.BytesIO(contents)
        wb = openpyxl.load_workbook(excel_stream, keep_links=True, data_only=False)

        total_rows, passed_rows, error_rows = process_check_C_logic(wb)

        output_stream = io.BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)

        filename_only = os.path.splitext(file.filename)[0]
        out_filename = f"{filename_only}_Checked_Python.xlsx"
        encoded_filename = quote(out_filename)

        headers = {
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "X-Total-Rows": str(total_rows),
            "X-Passed-Rows": str(passed_rows),
            "X-Error-Rows": str(error_rows)
        }

        return StreamingResponse(
            output_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"เกิดข้อผิดพลาดในการประมวลผลไฟล์: {str(e)}")
    
#----------------
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
